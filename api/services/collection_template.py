"""
Generate the user collection upload template (.xlsx) on demand.

Produces a workbook with three sheets:

* ``Collection`` -- the data sheet the user fills in. Frozen header,
  data validation dropdowns on Set / Condition / Is 1st Edition, number
  format on Quantity, currency format on Purchase Price.
* ``Instructions`` -- short usage notes plus a sample row.
* ``_lists`` -- hidden sheet that backs the Set dropdown via a named
  range. openpyxl's inline ``DataValidation.formula1`` strings are
  capped at 255 characters, so we put the full set list on its own
  sheet and reference it by range; this also lets the dropdown grow
  with the database without code changes.

The set list is queried fresh per request so newly added sets appear in
the dropdown without redeploying.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import text
from sqlalchemy.orm import Session


HEADERS = (
    ("Set", 32),
    ("Card Number", 14),
    ("Card Name", 30),
    ("Condition", 12),
    ("Variant", 24),
    ("Is 1st Edition", 16),
    ("Quantity", 12),
    ("Purchase Price", 16),
)
DATA_ROW_COUNT = 200  # apply data validation to this many rows of the template


def build_template_workbook(db: Session) -> bytes:
    """Return the bytes of a freshly generated upload template."""
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "Collection"

    _write_headers(data_ws)
    _apply_column_widths(data_ws)
    _apply_number_formats(data_ws)
    set_names = _query_set_names(db)
    _build_lists_sheet(wb, set_names)
    _apply_data_validations(wb, data_ws)
    _write_instructions_sheet(wb)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_filename(today: date | None = None) -> str:
    today = today or date.today()
    return f"card-collection-template-{today.isoformat()}.xlsx"


def _write_headers(ws) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2A2A3E")
    for idx, (label, _) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _apply_column_widths(ws) -> None:
    for idx, (_, width) in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _apply_number_formats(ws) -> None:
    quantity_col = _column_index("Quantity")
    price_col = _column_index("Purchase Price")
    for row in range(2, 2 + DATA_ROW_COUNT):
        ws.cell(row=row, column=quantity_col).number_format = "0"
        ws.cell(row=row, column=price_col).number_format = '"$"#,##0.00'


def _column_index(label: str) -> int:
    for idx, (header, _) in enumerate(HEADERS, start=1):
        if header == label:
            return idx
    raise KeyError(label)  # pragma: no cover -- HEADERS is a closed set


def _query_set_names(db: Session) -> list[str]:
    rows = db.execute(text("SELECT name FROM sets ORDER BY name")).fetchall()
    return [r[0] for r in rows if r[0]]


def _build_lists_sheet(wb: Workbook, set_names: list[str]) -> None:
    """Create a hidden sheet that holds the Set dropdown source list."""
    ws = wb.create_sheet("_lists")
    ws.sheet_state = "hidden"
    for idx, name in enumerate(set_names, start=1):
        ws.cell(row=idx, column=1, value=name)
    if set_names:
        last_row = len(set_names)
        ref = f"_lists!$A$1:$A${last_row}"
        defined = DefinedName(name="SetNames", attr_text=ref)
        wb.defined_names["SetNames"] = defined


def _apply_data_validations(wb: Workbook, ws) -> None:
    set_col = get_column_letter(_column_index("Set"))
    condition_col = get_column_letter(_column_index("Condition"))
    first_ed_col = get_column_letter(_column_index("Is 1st Edition"))

    if "SetNames" in wb.defined_names:
        set_dv = DataValidation(type="list", formula1="=SetNames", allow_blank=True)
        set_dv.error = "Pick a set from the dropdown."
        set_dv.errorTitle = "Unknown set"
        set_dv.add(f"{set_col}2:{set_col}{1 + DATA_ROW_COUNT}")
        ws.add_data_validation(set_dv)

    cond_dv = DataValidation(
        type="list",
        formula1='"NM,LP,MP,HP,DMG"',
        allow_blank=True,
    )
    cond_dv.error = "Use NM, LP, MP, HP, or DMG."
    cond_dv.errorTitle = "Unknown condition"
    cond_dv.add(f"{condition_col}2:{condition_col}{1 + DATA_ROW_COUNT}")
    ws.add_data_validation(cond_dv)

    bool_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    bool_dv.error = "TRUE or FALSE."
    bool_dv.errorTitle = "Invalid value"
    bool_dv.add(f"{first_ed_col}2:{first_ed_col}{1 + DATA_ROW_COUNT}")
    ws.add_data_validation(bool_dv)


def _write_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 110

    lines = [
        ("How to use this template", True),
        ("", False),
        ("1. Fill out one row per (card, condition, variant) combination you own.", False),
        ("2. Pick the Set, Condition, and Is 1st Edition values from their dropdowns.", False),
        ("3. Variant is free-form text. Type values like 'Reverse Holo' or 'Shadowless'.", False),
        ("   Multiple variants on one card can be separated with commas (e.g. 'Reverse Holo, Misprint').", False),
        ("4. Purchase Price is optional. Filling it in for every row enables the ROI KPIs in the dashboard.", False),
        ("5. Save the file and upload it back at /collection.", False),
        ("", False),
        ("Sample row:", True),
        ("   Set: Base Set | Card Number: 4 | Card Name: Charizard | Condition: NM | Variant: (blank) | Is 1st Edition: FALSE | Quantity: 1 | Purchase Price: 250.00", False),
        ("", False),
        ("Notes:", True),
        ("- Card Name is for your reference; the app matches on Set + Card Number.", False),
        ("- Extra columns (like an 'Error' column from a previous failed upload) are ignored on re-upload.", False),
        ("- Blank Is 1st Edition is treated as FALSE.", False),
    ]
    bold = Font(bold=True)
    for idx, (text_value, is_bold) in enumerate(lines, start=1):
        cell = ws.cell(row=idx, column=1, value=text_value)
        if is_bold:
            cell.font = bold
        cell.alignment = Alignment(wrap_text=True, vertical="top")
