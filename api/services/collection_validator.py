"""
Validate an uploaded collection workbook into structured rows.

Two failure modes:

* **Structural** -- a required header column is missing. The workbook is
  unusable as-is and the user must download a fresh template. We surface
  a single error string and skip row-level processing.
* **Row-level** -- the headers are present but individual rows fail one
  or more checks. The upload is rejected, but every offending row is
  collected so the frontend can build the annotated workbook.

The validator does not write to the database; it only resolves
``Set`` -> ``set_id`` and ``(set_id, card_number)`` -> ``cards.id``
through the session it is handed. Callers persist the result.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from schemas.collection import ParsedCollectionRow, RowError
from services.variant_normalizer import normalize as normalize_variant


REQUIRED_COLUMNS = (
    "Set",
    "Card Number",
    "Condition",
    "Is 1st Edition",
    "Quantity",
)
OPTIONAL_COLUMNS = ("Card Name", "Variant", "Purchase Price")
ALL_KNOWN_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

VALID_CONDITIONS = {"NM", "LP", "MP", "HP", "DMG"}
VALID_BOOLEANS = {"TRUE", "FALSE"}

# Header row in the template is row 1; data starts at row 2.
HEADER_ROW = 1
FIRST_DATA_ROW = 2

# The data sheet is the first sheet by index. Templates ship with sheet
# names "Collection" and "Instructions"; we tolerate any name on sheet 0
# so user-edited workbooks (rename, reorder) still parse.
DATA_SHEET_INDEX = 0


@dataclass
class ValidationResult:
    """Structured outcome of a single validation pass."""

    structural_error: str | None = None
    parsed_rows: list[ParsedCollectionRow] = field(default_factory=list)
    row_errors: list[RowError] = field(default_factory=list)
    total_rows: int = 0

    @property
    def has_errors(self) -> bool:
        return self.structural_error is not None or bool(self.row_errors)

    @property
    def distinct_error_messages(self) -> list[str]:
        seen: list[str] = []
        for err in self.row_errors:
            if err.message not in seen:
                seen.append(err.message)
        return seen


def validate_workbook(file_bytes: bytes, db: Session) -> ValidationResult:
    """Parse the bytes of an uploaded ``.xlsx`` and validate every row.

    Returns a :class:`ValidationResult` describing either the structural
    problem (header missing) or the per-row outcomes. The workbook is
    closed before returning so callers can re-open the same bytes for
    annotation without conflicting handles.
    """
    from io import BytesIO

    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    try:
        sheet = wb.worksheets[DATA_SHEET_INDEX]
        headers = _read_headers(sheet)
        missing = [c for c in REQUIRED_COLUMNS if c not in headers]
        if missing:
            return ValidationResult(
                structural_error=(
                    f"Required column '{missing[0]}' is missing -- "
                    "please use the latest template"
                )
            )

        set_lookup = _build_set_lookup(db)
        result = ValidationResult()
        # Iterate rows starting from FIRST_DATA_ROW; openpyxl rows are
        # 1-indexed and ws.iter_rows yields tuples of cells.
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True),
            start=FIRST_DATA_ROW,
        ):
            if _row_is_blank(row):
                continue
            result.total_rows += 1
            row_dict = _row_to_dict(headers, row)
            parsed, errors = _validate_one(row_dict, row_idx, set_lookup, db)
            if errors:
                result.row_errors.extend(errors)
            elif parsed is not None:
                result.parsed_rows.append(parsed)
        return result
    finally:
        wb.close()


def _read_headers(sheet) -> list[str]:
    """Return the header row as a list of stripped strings."""
    raw = next(sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    return [str(c).strip() if c is not None else "" for c in raw]


def _row_is_blank(row: tuple[Any, ...]) -> bool:
    return all(c is None or (isinstance(c, str) and c.strip() == "") for c in row)


def _row_to_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    """Pair headers with their cell values; ignores unknown columns later."""
    return dict(zip(headers, row))


def _build_set_lookup(db: Session) -> dict[str, str]:
    """Build a case-insensitive map from any known set label -> canonical set_id.

    Pulls every distinct identifier across set_identifiers plus the
    canonical id and display name from sets so the dropdown's display
    name and any historical synonym all resolve. The map keys are
    lowercased so callers index with ``label.strip().lower()``.
    """
    lookup: dict[str, str] = {}
    rows = db.execute(
        text(
            """
            SELECT s.id AS set_id, s.id AS label FROM sets s
            UNION
            SELECT s.id AS set_id, s.name AS label FROM sets s
            UNION
            SELECT si.set_id, si.identifier AS label FROM set_identifiers si
            """
        )
    ).fetchall()
    for row in rows:
        if row.label is None:
            continue
        lookup[str(row.label).strip().lower()] = row.set_id
    return lookup


def _validate_one(
    row: dict[str, Any],
    row_number: int,
    set_lookup: dict[str, str],
    db: Session,
) -> tuple[ParsedCollectionRow | None, list[RowError]]:
    """Validate one workbook row and return either the parsed row or errors."""
    errors: list[RowError] = []

    # Set -> set_id.
    raw_set = row.get("Set")
    set_id: str | None = None
    if raw_set is None or str(raw_set).strip() == "":
        errors.append(RowError(row_number=row_number, message="Set is required"))
    else:
        key = str(raw_set).strip().lower()
        set_id = set_lookup.get(key)
        if set_id is None:
            errors.append(
                RowError(
                    row_number=row_number,
                    message=f"Set '{raw_set}' is not recognized",
                )
            )

    # Card Number (Excel often gives floats like 4.0 -> coerce to a string).
    raw_number = row.get("Card Number")
    card_number_text: str | None = None
    if raw_number is None or str(raw_number).strip() == "":
        errors.append(RowError(row_number=row_number, message="Card Number is required"))
    else:
        card_number_text = _coerce_number_to_text(raw_number)

    # Card resolution requires both a set and a card number to have validated.
    card_id: str | None = None
    if set_id is not None and card_number_text is not None:
        card_id = _resolve_card_id(db, set_id, card_number_text)
        if card_id is None:
            errors.append(
                RowError(
                    row_number=row_number,
                    message=(
                        f"Card number {card_number_text} does not exist in {raw_set}"
                    ),
                )
            )

    # Condition.
    raw_condition = row.get("Condition")
    condition: str | None = None
    if raw_condition is None or str(raw_condition).strip() == "":
        errors.append(RowError(row_number=row_number, message="Condition is required"))
    else:
        candidate = str(raw_condition).strip().upper()
        if candidate not in VALID_CONDITIONS:
            errors.append(
                RowError(
                    row_number=row_number,
                    message="Condition must be NM, LP, MP, HP, or DMG",
                )
            )
        else:
            condition = candidate

    # Is 1st Edition (blank treated as FALSE).
    raw_first = row.get("Is 1st Edition")
    is_first: bool = False
    if raw_first is None or str(raw_first).strip() == "":
        is_first = False
    else:
        candidate = str(raw_first).strip().upper()
        if candidate not in VALID_BOOLEANS:
            errors.append(
                RowError(
                    row_number=row_number,
                    message="Is 1st Edition must be TRUE or FALSE",
                )
            )
        else:
            is_first = candidate == "TRUE"

    # Quantity.
    raw_quantity = row.get("Quantity")
    quantity: int | None = None
    if raw_quantity is None or str(raw_quantity).strip() == "":
        errors.append(RowError(row_number=row_number, message="Quantity is required"))
    else:
        coerced = _coerce_int(raw_quantity)
        if coerced is None or coerced < 1:
            errors.append(
                RowError(
                    row_number=row_number,
                    message="Quantity must be a whole number greater than 0",
                )
            )
        else:
            quantity = coerced

    # Purchase Price (optional).
    raw_price = row.get("Purchase Price")
    purchase_price: Decimal | None = None
    if raw_price is not None and str(raw_price).strip() != "":
        try:
            purchase_price = Decimal(str(raw_price).strip().replace("$", "").replace(",", ""))
        except (InvalidOperation, ValueError):
            errors.append(
                RowError(
                    row_number=row_number,
                    message="Purchase Price must be a number or left blank",
                )
            )

    # Variant (free-form, normalized in process).
    variants = normalize_variant(row.get("Variant"))

    if errors:
        return None, errors
    return (
        ParsedCollectionRow(
            card_id=card_id,  # type: ignore[arg-type]
            condition=condition,  # type: ignore[arg-type]
            variant=variants,
            is_first_edition=is_first,
            quantity=quantity,  # type: ignore[arg-type]
            purchase_price=purchase_price,
        ),
        [],
    )


def _coerce_number_to_text(value: Any) -> str:
    """Render a number like ``4`` or ``4.0`` as ``"4"`` while leaving alphanumeric strings alone."""
    if isinstance(value, bool):
        # bool subclasses int -- treat as plain string.
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _coerce_int(value: Any) -> int | None:
    """Accept ints and integer-valued floats; reject anything else."""
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None
    try:
        text_val = str(value).strip()
        as_float = float(text_val)
    except (TypeError, ValueError):
        return None
    if as_float.is_integer():
        return int(as_float)
    return None


def _resolve_card_id(db: Session, set_id: str, card_number: str) -> str | None:
    row = db.execute(
        text(
            """
            SELECT id FROM cards
            WHERE set_id = :set_id AND number = :number
            LIMIT 1
            """
        ),
        {"set_id": set_id, "number": card_number},
    ).fetchone()
    return row[0] if row else None
