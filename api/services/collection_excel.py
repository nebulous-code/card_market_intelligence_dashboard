"""
Populate the collection Excel template with the user's collection data.

The user clicks "Download Excel" on the dashboard. The endpoint reads
the session, asks this service to fill the template, and streams the
bytes back as an attachment. The user opens the file and triggers
"Refresh All" -- charts and pivots in the real (post-implementation)
template re-derive from the four tables we populated here.

Four tables are populated, each by name:

* ``collection_details`` -- one row per session line item.
* ``condition_multipliers`` -- multiplier rows filtered to user's sets.
* ``historic_prices`` -- twice-monthly snapshots over the last 6 months.
* ``card_prices_all_conditions`` -- current prices at upgrade conditions.

The endpoint always returns the user's full collection. Slicer state on
the dashboard is intentionally ignored -- the Excel is a portable copy,
not a snapshot of the current view.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session

from schemas.collection import ParsedCollectionRow


TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "assets" / "collection_template.xlsx"
)

# Raw conditions that have full snapshot coverage. Anything else
# (PSA-10, BGS-9.5, unknown values) is filtered out of the historic
# prices and card_prices tables -- their data lives elsewhere or not at
# all.
RAW_CONDITIONS = ("NM", "LP", "MP", "HP", "DMG")

# Upgrade ladder: for a card at condition X, the conditions to look at
# for the upgrade-cost analyzer. The dictionary's key order matters for
# the response -- we list nicer-than-X conditions, not X itself.
UPGRADE_TARGETS: dict[str, tuple[str, ...]] = {
    "DMG": ("HP", "MP", "LP", "NM"),
    "HP": ("MP", "LP", "NM"),
    "MP": ("LP", "NM"),
    "LP": ("NM",),
    "NM": (),
}

# How many months of history to sample. Spec says 6 months, twice
# monthly (1st and 15th).
HISTORY_MONTHS = 6
SAMPLE_DAYS = (1, 15)


# -----------------------------------------------------------------------------
# Public API
# -----------------------------------------------------------------------------


def populate_template(
    db: Session, parsed_rows: list[ParsedCollectionRow]
) -> bytes:
    """Open the template, write the four tables, return the bytes.

    Raises ``FileNotFoundError`` if the template asset is missing on
    the deployment -- the router catches this and returns 503.
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(str(TEMPLATE_PATH))

    detail_rows = _collect_details_rows(db, parsed_rows)
    user_set_ids = sorted({row["set_id"] for row in detail_rows})
    multiplier_rows = _collect_multipliers_rows(db, user_set_ids)
    historic_rows = _collect_historic_rows(db, parsed_rows)
    upgrade_rows = _collect_upgrade_rows(db, parsed_rows)

    # ``keep_links=True`` preserves any external references the real
    # template may rely on (Power Query connections live as
    # connections.xml, which openpyxl carries through unmodified).
    wb = load_workbook(filename=TEMPLATE_PATH, keep_vba=False, keep_links=True)
    try:
        _write_table(wb, "collection_details", _DETAILS_COLUMNS, detail_rows)
        _write_table(wb, "condition_multipliers", _MULTIPLIERS_COLUMNS, multiplier_rows)
        _write_table(wb, "historic_prices", _HISTORIC_COLUMNS, historic_rows)
        _write_table(wb, "card_prices_all_conditions", _UPGRADE_COLUMNS, upgrade_rows)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()


# -----------------------------------------------------------------------------
# Row collection -- one helper per table, all pure-ish (single SQL call each)
# -----------------------------------------------------------------------------


def _collect_details_rows(
    db: Session, parsed_rows: list[ParsedCollectionRow]
) -> list[dict]:
    """One row per session line item with computed value / gain fields."""
    if not parsed_rows:
        return []

    card_ids = sorted({r.card_id for r in parsed_rows})
    metadata = _fetch_card_metadata(db, card_ids)
    conditions_needed = sorted({r.condition for r in parsed_rows} | {"NM"})
    prices = _fetch_latest_prices(db, card_ids, conditions_needed)

    out: list[dict] = []
    for row in parsed_rows:
        meta = metadata.get(row.card_id)
        if meta is None:  # pragma: no cover -- session rows reference live cards
            continue
        market_price, fellback = _resolve_market_price(prices, row.card_id, row.condition)
        has_variant = bool(row.variant)
        pricing_warning = bool(fellback or has_variant)

        quantity = row.quantity
        purchase_total = (
            row.purchase_price * quantity if row.purchase_price is not None else None
        )
        if market_price is not None:
            total_value: Decimal | None = market_price * quantity
        else:
            total_value = None
        if purchase_total is not None and total_value is not None:
            gain_dollar: Decimal | None = total_value - purchase_total
            gain_percent: Decimal | None = (
                (gain_dollar / purchase_total) if purchase_total != 0 else None
            )
        else:
            gain_dollar = None
            gain_percent = None

        out.append(
            {
                "set_id": meta["set_id"],
                "set_name": meta["set_name"],
                "set_printed_total": meta["printed_total"],
                # ``total_count`` is the live COUNT of cards in the set
                # (M03_S07's substitution for ``total_with_secrets``);
                # they cover the same intent, so we map it here.
                "set_total_with_secrets": meta["total_count"],
                "card_id": row.card_id,
                "card_number": meta["card_number"],
                "card_name": meta["name"],
                "rarity": meta["rarity"],
                "supertype": meta["supertype"],
                "condition": row.condition,
                "variant": ", ".join(row.variant) if row.variant else None,
                "is_first_edition": bool(row.is_first_edition),
                "quantity": quantity,
                "purchase_price": row.purchase_price,
                "market_price": market_price,
                "total_value": total_value,
                "gain_dollar": gain_dollar,
                "gain_percent": gain_percent,
                "pricing_warning": pricing_warning,
            }
        )
    return out


def _collect_multipliers_rows(db: Session, set_ids: list[str]) -> list[dict]:
    """All multiplier rows for the user's sets, joined to set name."""
    if not set_ids:
        return []
    rows = db.execute(
        text(
            """
            SELECT
                cm.set_id        AS set_id,
                s.name           AS set_name,
                cm.grouping_type AS grouping_type,
                cm.grouping_value AS grouping_value,
                cm.from_condition AS from_condition,
                cm.to_condition   AS to_condition,
                cm.multiplier     AS multiplier,
                cm.data_points    AS data_points
            FROM condition_multipliers cm
            JOIN sets s ON s.id = cm.set_id
            WHERE cm.set_id = ANY(:ids)
            ORDER BY cm.set_id, cm.grouping_type, cm.grouping_value,
                     cm.from_condition, cm.to_condition
            """
        ),
        {"ids": set_ids},
    ).fetchall()
    return [
        {
            "set_id": r.set_id,
            "set_name": r.set_name,
            "grouping_type": r.grouping_type,
            "grouping_value": r.grouping_value,
            "from_condition": r.from_condition,
            "to_condition": r.to_condition,
            "multiplier": r.multiplier,
            "data_points": r.data_points,
        }
        for r in rows
    ]


def _collect_historic_rows(
    db: Session, parsed_rows: list[ParsedCollectionRow]
) -> list[dict]:
    """Twice-monthly snapshots for each card+condition over 6 months.

    Cards on graded / unrecognized conditions contribute no rows. Sample
    dates with no qualifying snapshot are skipped (no NULL prices).
    """
    keys = sorted(
        {(r.card_id, r.condition) for r in parsed_rows if r.condition in RAW_CONDITIONS}
    )
    if not keys:
        return []

    card_ids = sorted({k[0] for k in keys})
    conditions = sorted({k[1] for k in keys})
    metadata = _fetch_card_metadata(db, card_ids)

    # Pull every snapshot for these (card, condition) pairs in one query;
    # filter and bucket per-key in Python so we only round-trip once.
    rows = db.execute(
        text(
            """
            SELECT DISTINCT ON (card_id, condition, captured_date)
                card_id, condition, captured_date, market_price
            FROM price_snapshots
            WHERE card_id = ANY(:card_ids)
              AND condition = ANY(:conditions)
              AND source = 'tcgplayer'
              AND market_price IS NOT NULL
            ORDER BY
                card_id,
                condition,
                captured_date,
                (variant IS NULL) DESC,
                captured_at
            """
        ),
        {"card_ids": card_ids, "conditions": conditions},
    ).fetchall()

    by_key: dict[tuple[str, str], list[_Snap]] = defaultdict(list)
    keys_set = set(keys)
    for r in rows:
        key = (r.card_id, r.condition)
        if key not in keys_set:
            continue  # pragma: no cover -- ANY/ANY filter is exact
        by_key[key].append(_Snap(r.captured_date, r.market_price))

    sample_dates = _sample_dates(date.today(), HISTORY_MONTHS)
    out: list[dict] = []
    for key in keys:
        snaps = by_key.get(key, [])
        if not snaps:
            continue
        meta = metadata.get(key[0])
        if meta is None:  # pragma: no cover
            continue
        for sample in sample_dates:
            price = _locf_price(snaps, sample)
            if price is None:
                continue
            out.append(
                {
                    "card_id": key[0],
                    "card_name": meta["name"],
                    "set_name": meta["set_name"],
                    "condition": key[1],
                    "sample_date": sample,
                    "market_price": price,
                }
            )
    return out


def _collect_upgrade_rows(
    db: Session, parsed_rows: list[ParsedCollectionRow]
) -> list[dict]:
    """Current prices at conditions above the user's, for raw cards only."""
    keys = sorted(
        {
            (r.card_id, r.condition)
            for r in parsed_rows
            if r.condition in UPGRADE_TARGETS and UPGRADE_TARGETS[r.condition]
        }
    )
    if not keys:
        return []

    card_ids = sorted({k[0] for k in keys})
    targets_needed = sorted(
        {target for _, cond in keys for target in UPGRADE_TARGETS[cond]}
    )
    metadata = _fetch_card_metadata(db, card_ids)
    prices = _fetch_latest_prices(db, card_ids, targets_needed)

    out: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for card_id, condition in keys:
        meta = metadata.get(card_id)
        if meta is None:  # pragma: no cover
            continue
        for target in UPGRADE_TARGETS[condition]:
            dedup_key = (card_id, target)
            if dedup_key in seen:
                continue
            seen.add(dedup_key)
            market_price = prices.get((card_id, target))
            if market_price is None:
                continue
            out.append(
                {
                    "card_id": card_id,
                    "card_name": meta["name"],
                    "set_name": meta["set_name"],
                    "condition": target,
                    "market_price": market_price,
                }
            )
    return out


# -----------------------------------------------------------------------------
# Workbook writing
# -----------------------------------------------------------------------------


_DETAILS_COLUMNS = (
    "set_id",
    "set_name",
    "set_printed_total",
    "set_total_with_secrets",
    "card_id",
    "card_number",
    "card_name",
    "rarity",
    "supertype",
    "condition",
    "variant",
    "is_first_edition",
    "quantity",
    "purchase_price",
    "market_price",
    "total_value",
    "gain_dollar",
    "gain_percent",
    "pricing_warning",
)
_MULTIPLIERS_COLUMNS = (
    "set_id",
    "set_name",
    "grouping_type",
    "grouping_value",
    "from_condition",
    "to_condition",
    "multiplier",
    "data_points",
)
_HISTORIC_COLUMNS = (
    "card_id",
    "card_name",
    "set_name",
    "condition",
    "sample_date",
    "market_price",
)
_UPGRADE_COLUMNS = (
    "card_id",
    "card_name",
    "set_name",
    "condition",
    "market_price",
)


def _write_table(wb, table_name: str, columns: tuple[str, ...], rows: list[dict]) -> None:
    """Append ``rows`` to the Table named ``table_name``.

    Existing data rows are cleared first; the header row is preserved
    in place. The Table's ``ref`` is updated so Excel recognizes the
    new last row.
    """
    sheet, table = _find_table(wb, table_name)
    _clear_existing_data_rows(sheet, table, columns)

    last_col_letter = get_column_letter(len(columns))
    if not rows:
        # Excel requires Tables to span >= 2 rows. Keep the original
        # placeholder data row (one blank row beneath the header) so
        # the file opens without an inconsistency warning.
        table.ref = f"A1:{last_col_letter}2"
        return

    first_data_row = 2
    last_data_row = first_data_row + len(rows) - 1
    for offset, row in enumerate(rows):
        row_idx = first_data_row + offset
        for col_idx, name in enumerate(columns, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=_excel_value(row.get(name)))

    table.ref = f"A1:{last_col_letter}{last_data_row}"


def _find_table(wb, table_name: str):
    for sheet in wb.worksheets:
        if table_name in sheet.tables:
            return sheet, sheet.tables[table_name]
    raise KeyError(f"Table {table_name!r} not found in template")  # pragma: no cover


def _clear_existing_data_rows(sheet, table, columns: tuple[str, ...]) -> None:
    """Wipe any rows under the header so populated rows don't mix."""
    if not table.ref:  # pragma: no cover -- generated templates always have a ref
        return
    # ``ref`` is e.g. "A1:S2"; pull the trailing row number.
    end = table.ref.split(":")[1]
    last_row = int("".join(ch for ch in end if ch.isdigit()))
    if last_row < 2:
        return  # pragma: no cover -- header-only refs aren't generated
    for row_idx in range(2, last_row + 1):
        for col_idx in range(1, len(columns) + 1):
            sheet.cell(row=row_idx, column=col_idx, value=None)


def _excel_value(value: Any) -> Any:
    """Coerce values to types Excel handles natively.

    Decimal is converted to float so cells display as numbers (the
    default openpyxl behaviour stringifies Decimals). ``None`` stays
    ``None`` so cells are blank for ``ISBLANK`` checks.
    """
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return value


# -----------------------------------------------------------------------------
# SQL helpers
# -----------------------------------------------------------------------------


@dataclass(frozen=True)
class _Snap:
    captured_date: date
    market_price: Decimal


def _fetch_card_metadata(db: Session, card_ids: list[str]) -> dict[str, dict]:
    """Card + set metadata keyed by card_id."""
    if not card_ids:  # pragma: no cover -- callers short-circuit on empty input
        return {}
    rows = db.execute(
        text(
            """
            SELECT
                c.id            AS card_id,
                c.name          AS name,
                c.number        AS card_number,
                c.rarity        AS rarity,
                c.supertype     AS supertype,
                c.set_id        AS set_id,
                s.name          AS set_name,
                s.printed_total AS printed_total,
                (SELECT COUNT(*) FROM cards c2 WHERE c2.set_id = s.id) AS total_count
            FROM cards c
            JOIN sets s ON s.id = c.set_id
            WHERE c.id = ANY(:ids)
            """
        ),
        {"ids": card_ids},
    ).fetchall()
    return {
        r.card_id: {
            "name": r.name,
            "card_number": r.card_number,
            "rarity": r.rarity,
            "supertype": r.supertype,
            "set_id": r.set_id,
            "set_name": r.set_name,
            "printed_total": r.printed_total,
            "total_count": r.total_count,
        }
        for r in rows
    }


def _fetch_latest_prices(
    db: Session, card_ids: list[str], conditions: list[str]
) -> dict[tuple[str, str], Decimal]:
    """Latest price per ``(card_id, condition)``, NULL-variant preferred."""
    if not card_ids or not conditions:  # pragma: no cover -- callers guard
        return {}
    rows = db.execute(
        text(
            """
            SELECT DISTINCT ON (card_id, condition)
                card_id, condition, market_price
            FROM price_snapshots
            WHERE card_id = ANY(:ids)
              AND condition = ANY(:conditions)
              AND source = 'tcgplayer'
              AND market_price IS NOT NULL
            ORDER BY
                card_id,
                condition,
                captured_at DESC,
                (variant IS NULL) DESC,
                variant
            """
        ),
        {"ids": card_ids, "conditions": conditions},
    ).fetchall()
    return {(r.card_id, r.condition): r.market_price for r in rows}


def _resolve_market_price(
    prices: dict[tuple[str, str], Decimal], card_id: str, condition: str
) -> tuple[Decimal | None, bool]:
    """Return ``(price, fellback_to_nm)`` for the given key.

    Falls back to NM market price when no snapshot exists for the
    user's exact condition, signalling ``pricing_warning`` upstream.
    """
    direct = prices.get((card_id, condition))
    if direct is not None:
        return direct, False
    if condition != "NM":
        nm_price = prices.get((card_id, "NM"))
        if nm_price is not None:
            return nm_price, True
    return None, False


def _sample_dates(today: date, months: int) -> list[date]:
    """Twice-monthly sample anchors ascending up to today.

    Returns the 1st + 15th of each of the past ``months`` months that
    falls on or before ``today``. Future-dated anchors (e.g. the 15th
    of the current month before the 15th has arrived) are skipped.
    """
    dates: list[date] = []
    # Start from the 1st of the month that's ``months - 1`` months before
    # the current month to capture roughly ``months`` months of history.
    cursor_year = today.year
    cursor_month = today.month - (months - 1)
    while cursor_month <= 0:
        cursor_month += 12
        cursor_year -= 1
    for _ in range(months):
        for day in SAMPLE_DAYS:
            anchor = date(cursor_year, cursor_month, day)
            if anchor <= today:
                dates.append(anchor)
        cursor_month += 1
        if cursor_month > 12:
            cursor_month = 1
            cursor_year += 1
    return dates


def _locf_price(snaps: list[_Snap], target: date) -> Decimal | None:
    """Most recent snapshot at or before ``target``; ``None`` if no
    qualifying snapshot exists. Unlike the dashboard timeseries we
    explicitly do NOT fall back to the earliest known price -- the
    spec says skip the row entirely when no data exists at the sample
    date."""
    chosen: Decimal | None = None
    for s in snaps:
        if s.captured_date <= target:
            chosen = s.market_price
        else:
            break
    return chosen


def excel_filename(today: date | None = None) -> str:
    today = today or date.today()
    return f"collection-report-{today.isoformat()}.xlsx"


# Quiet "imported but unused" complaints for symbols re-exported by
# default through ``from .collection_excel import *``.
__all__ = [
    "EXCEL_MEDIA_TYPE",
    "TEMPLATE_PATH",
    "excel_filename",
    "populate_template",
]


EXCEL_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
