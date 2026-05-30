"""
Annotate a failed-upload workbook with per-row error messages.

The user uploads ``foo.xlsx``, validation rejects some rows, and the
frontend offers a "Download Annotated Workbook" button. That button hits
``POST /collection/upload/annotated`` with the original file. This
service re-runs validation server-side and produces the annotated copy:

* Adds (or reuses) an ``Error`` column on the right edge of the data
  sheet. Header is bold red.
* Writes the row's error message into the Error column for invalid rows.
  Valid rows leave the cell blank, so re-uploading the same file after a
  fix succeeds without the user having to delete the column.
* Fills invalid rows with ``#5C2828``.

Re-validation is intentional: round-tripping the previous validation
result through the client would be brittle (large, easy to mismatch,
and easy to forge). Same code path, no extra wire format.
"""

from __future__ import annotations

from collections import defaultdict
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from services.collection_validator import (
    DATA_SHEET_INDEX,
    FIRST_DATA_ROW,
    HEADER_ROW,
    validate_workbook,
)

ERROR_HEADER = "Error"
ERROR_FILL = PatternFill("solid", fgColor="5C2828")


def annotate_workbook(file_bytes: bytes, db: Session) -> bytes:
    """Return bytes of the input workbook with errors marked.

    If the workbook has a structural problem we still return the
    original file with a single ``Error`` cell on row 2 explaining the
    issue. That way the user always sees a meaningful annotated file
    download even when the structural error came from deleting a header.
    """
    result = validate_workbook(file_bytes, db)

    wb = load_workbook(filename=BytesIO(file_bytes))
    try:
        sheet = wb.worksheets[DATA_SHEET_INDEX]
        error_col = _ensure_error_column(sheet)

        if result.structural_error:
            sheet.cell(row=FIRST_DATA_ROW, column=error_col, value=result.structural_error)
            return _serialize(wb)

        # Group errors by row so multi-error rows render their messages joined.
        by_row: dict[int, list[str]] = defaultdict(list)
        for err in result.row_errors:
            by_row[err.row_number].append(err.message)

        for row_idx, messages in by_row.items():
            sheet.cell(row=row_idx, column=error_col, value="; ".join(messages))
            for col_idx in range(1, error_col + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = ERROR_FILL

        return _serialize(wb)
    finally:
        wb.close()


def _ensure_error_column(sheet) -> int:
    """Locate or create the Error column; return its 1-based index."""
    headers = next(
        sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True)
    )
    for idx, value in enumerate(headers, start=1):
        if value is not None and str(value).strip() == ERROR_HEADER:
            return idx
    new_col = (sheet.max_column or len(headers)) + 1
    cell = sheet.cell(row=HEADER_ROW, column=new_col, value=ERROR_HEADER)
    cell.font = Font(bold=True, color="C53030")
    sheet.column_dimensions[get_column_letter(new_col)].width = 60
    return new_col


def _serialize(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
