"""
Generate ``api/assets/mock_collection.xlsx`` from currently-ingested data.

Run once locally after a database with at least Base Set / Jungle /
Fossil / Pokemon 151 ingested:

    cd api && python -m scripts.generate_mock_collection

The resulting file is checked into the repo so the ``/collection/mock``
endpoint has something to load on a clean deploy. The script only needs
to be re-run when the database shape changes meaningfully (e.g. a set
gets renamed or removed); a stale mock will just fail validation, which
is loud and easy to diagnose.

The mock contains 20 rows pulled from up to four sets, mixes the five
conditions, leaves variants blank for simplicity, and provides purchase
prices on roughly half the rows so the dashboard can exercise both the
with-purchase and without-purchase code paths.
"""

from __future__ import annotations

import sys
from itertools import cycle
from pathlib import Path

# Make ``database`` and friends importable when this is run via
# ``python -m scripts.generate_mock_collection`` from the api/ directory.
HERE = Path(__file__).resolve()
API_ROOT = HERE.parent.parent
if str(API_ROOT) not in sys.path:
    sys.path.insert(0, str(API_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text

from database import SessionLocal


PREFERRED_SETS = ("base1", "base2", "base3", "sv03.5")
ROWS_PER_SET = 5
CONDITIONS = ("NM", "NM", "LP", "MP", "HP", "DMG")
# Mix of blank, single, and multi-variant entries plus a couple of
# 1st-edition rows so the dashboard's variant chart and 1st-Ed bucket
# are exercised by the bundled mock.
VARIANTS = (
    None,
    None,
    "Reverse Holo",
    None,
    "Holo",
    None,
    "Reverse Holo, Misprint",
    None,
    "Shadowless",
    None,
)
FIRST_EDITION_FLAGS = (
    "FALSE",
    "FALSE",
    "FALSE",
    "FALSE",
    "TRUE",
    "FALSE",
    "FALSE",
    "TRUE",
    "FALSE",
    "FALSE",
)
PURCHASE_PRICE_TEMPLATE = (
    250.00,
    180.00,
    12.50,
    None,
    5.00,
    320.00,
    None,
    25.00,
    8.50,
    None,
)
OUTPUT = API_ROOT / "assets" / "mock_collection.xlsx"


def main() -> None:
    db = SessionLocal()
    try:
        rows = _gather_rows(db)
    finally:
        db.close()

    if not rows:
        raise SystemExit(
            "No cards found for the preferred sets. Ingest at least one set "
            "before generating the mock collection."
        )

    wb = _build_workbook(rows)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {len(rows)} rows -> {OUTPUT}")


def _gather_rows(db) -> list[dict]:
    out: list[dict] = []
    cond_cycle = cycle(CONDITIONS)
    price_cycle = cycle(PURCHASE_PRICE_TEMPLATE)
    variant_cycle = cycle(VARIANTS)
    first_ed_cycle = cycle(FIRST_EDITION_FLAGS)
    for set_id in PREFERRED_SETS:
        info = db.execute(
            text("SELECT id, name FROM sets WHERE id = :id"),
            {"id": set_id},
        ).fetchone()
        if info is None:
            continue
        cards = db.execute(
            text(
                """
                SELECT id, number, name FROM cards
                WHERE set_id = :id
                ORDER BY length(number), number
                LIMIT :limit
                """
            ),
            {"id": set_id, "limit": ROWS_PER_SET},
        ).fetchall()
        for card in cards:
            out.append(
                {
                    "set_name": info.name,
                    "card_number": card.number,
                    "card_name": card.name,
                    "condition": next(cond_cycle),
                    "variant": next(variant_cycle),
                    "is_first_edition": next(first_ed_cycle),
                    "quantity": 1,
                    "purchase_price": next(price_cycle),
                }
            )
    return out


def _build_workbook(rows: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Collection"
    headers = [
        "Set",
        "Card Number",
        "Card Name",
        "Condition",
        "Variant",
        "Is 1st Edition",
        "Quantity",
        "Purchase Price",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2A2A3E")
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, entry in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=entry["set_name"])
        ws.cell(row=row_idx, column=2, value=entry["card_number"])
        ws.cell(row=row_idx, column=3, value=entry["card_name"])
        ws.cell(row=row_idx, column=4, value=entry["condition"])
        ws.cell(row=row_idx, column=5, value=entry["variant"])
        ws.cell(row=row_idx, column=6, value=entry["is_first_edition"])
        ws.cell(row=row_idx, column=7, value=entry["quantity"])
        ws.cell(row=row_idx, column=8, value=entry["purchase_price"])
    return wb


if __name__ == "__main__":
    main()
