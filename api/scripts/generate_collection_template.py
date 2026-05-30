"""
Generate the placeholder ``api/assets/collection_template.xlsx`` workbook.

Run once to seed the template the populator will fill at runtime:

    cd api && python -m scripts.generate_collection_template

The result has four sheets each holding one named Excel ``Table``
(ListObject) with headers only. The populator opens this file, appends
collection-derived rows to the four tables, and returns the result as
a download. The real template -- with Power Query connections, pivots,
and charts -- is built outside the agent's scope and replaces this
file when ready (see M04_S05).

This generator is operator tooling, not application code. It is in
``api/scripts/`` (already excluded from coverage) so it does not need
test scaffolding.
"""

from __future__ import annotations

import sys
from pathlib import Path

HERE = Path(__file__).resolve()
API_ROOT = HERE.parent.parent
if str(API_ROOT) not in sys.path:
    sys.path.insert(0, str(API_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------- table definitions ----------
#
# Edit these to add or rename columns. The populator imports the same
# constants so headers stay in sync between generator and runtime.

COLLECTION_DETAILS_HEADERS = [
    "set_id",
    "set_name",
    "set_printed_total",
    "set_total_with_secrets",
    "card_id",
    "card_number",
    "card_name",
    "rarity",
    "supertype",
    "condition",
    "variant",
    "is_first_edition",
    "quantity",
    "purchase_price",
    "market_price",
    "total_value",
    "gain_dollar",
    "gain_percent",
    "pricing_warning",
]

CONDITION_MULTIPLIERS_HEADERS = [
    "set_id",
    "set_name",
    "grouping_type",
    "grouping_value",
    "from_condition",
    "to_condition",
    "multiplier",
    "data_points",
]

HISTORIC_PRICES_HEADERS = [
    "card_id",
    "card_name",
    "set_name",
    "condition",
    "sample_date",
    "market_price",
]

CARD_PRICES_HEADERS = [
    "card_id",
    "card_name",
    "set_name",
    "condition",
    "market_price",
]


SHEETS = [
    ("Collection Details", "collection_details", COLLECTION_DETAILS_HEADERS),
    ("Condition Multipliers", "condition_multipliers", CONDITION_MULTIPLIERS_HEADERS),
    ("Historic Prices", "historic_prices", HISTORIC_PRICES_HEADERS),
    ("Card Prices", "card_prices_all_conditions", CARD_PRICES_HEADERS),
]


OUTPUT = API_ROOT / "assets" / "collection_template.xlsx"


def main() -> None:
    wb = Workbook()
    # Drop the default sheet that openpyxl creates on Workbook().
    wb.remove(wb.active)

    for sheet_name, table_name, headers in SHEETS:
        ws = wb.create_sheet(sheet_name)
        _write_headers(ws, headers)
        _add_table(ws, table_name, headers)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} with {len(SHEETS)} sheets and tables.")


def _write_headers(ws, headers: list[str]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2A2A3E")
    for idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
    # Sensible default column widths -- the real template overrides these.
    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    ws.freeze_panes = "A2"


def _add_table(ws, table_name: str, headers: list[str]) -> None:
    """Add a named Excel Table covering header row + one placeholder row.

    Excel requires Tables to span at least one data row in addition to
    the header. We seed an empty data row so the file opens cleanly;
    the populator overwrites or removes that row when filling the
    table at runtime.
    """
    last_col = get_column_letter(len(headers))
    ref = f"A1:{last_col}2"  # header row + 1 placeholder data row
    table = Table(displayName=table_name, name=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


if __name__ == "__main__":
    main()
