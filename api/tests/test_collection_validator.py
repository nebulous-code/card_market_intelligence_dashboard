"""
Tests for services.collection_validator.

The validator depends on a database session for two lookups:
``set_identifiers`` and ``cards``. Tests reuse the existing
``sample_set`` / ``sample_cards`` fixtures and add the bare
``set_identifiers`` rows that production seed data fills in via
migration 005 (which only seeds Base Set; the test DB also has Base Set
plus whatever the per-test fixture adds).

Each test builds an in-memory workbook with openpyxl rather than
reading a file, so the tests are deterministic and exercise the column
ordering / blank-row / coercion logic directly.
"""

from io import BytesIO

from openpyxl import Workbook

from services.collection_validator import (
    REQUIRED_COLUMNS,
    ValidationResult,
    validate_workbook,
)


HEADERS = (
    "Set",
    "Card Number",
    "Card Name",
    "Condition",
    "Variant",
    "Is 1st Edition",
    "Quantity",
    "Purchase Price",
)


def _build_workbook(rows: list[dict] | None = None, headers: tuple[str, ...] = HEADERS) -> bytes:
    """Helper: write a workbook with the given headers and data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Collection"
    for col_idx, label in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=label)
    if rows:
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, label in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(label))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _valid_row(**overrides):
    base = {
        "Set": "Base Set",
        "Card Number": 4,
        "Card Name": "Charizard",
        "Condition": "NM",
        "Variant": None,
        "Is 1st Edition": "FALSE",
        "Quantity": 1,
        "Purchase Price": None,
    }
    base.update(overrides)
    return base


def test_has_errors_default_is_false():
    """Empty result -> no errors, no structural problem."""
    assert ValidationResult().has_errors is False


def test_has_errors_true_when_structural():
    assert ValidationResult(structural_error="missing").has_errors is True


def test_distinct_error_messages_dedupes_preserving_order():
    """Distinct messages list is order-stable and removes duplicates."""
    from schemas.collection import RowError

    result = ValidationResult(
        row_errors=[
            RowError(row_number=2, message="bad set"),
            RowError(row_number=3, message="bad number"),
            RowError(row_number=4, message="bad set"),
        ],
    )
    assert result.distinct_error_messages == ["bad set", "bad number"]


def test_missing_required_column_is_structural(client, db_session, sample_cards):
    headers = tuple(h for h in HEADERS if h != "Condition")
    blob = _build_workbook(rows=[_valid_row()], headers=headers)
    result = validate_workbook(blob, db_session)
    assert result.structural_error is not None
    assert "Condition" in result.structural_error
    assert result.has_errors is True
    # Row-level processing is skipped on structural failure.
    assert result.parsed_rows == []
    assert result.row_errors == []


def test_blank_rows_are_ignored(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(), {}, _valid_row()])
    result = validate_workbook(blob, db_session)
    assert result.total_rows == 2
    assert len(result.parsed_rows) == 2


def test_valid_row_parses(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row()])
    result = validate_workbook(blob, db_session)
    assert result.has_errors is False
    assert len(result.parsed_rows) == 1
    parsed = result.parsed_rows[0]
    assert parsed.card_id == "base1-4"
    assert parsed.condition == "NM"
    assert parsed.variant == []
    assert parsed.is_first_edition is False
    assert parsed.quantity == 1
    assert parsed.purchase_price is None


def test_set_resolves_via_canonical_id(db_session, sample_cards):
    """``base1`` works just as well as ``Base Set``."""
    blob = _build_workbook(rows=[_valid_row(Set="base1")])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors
    assert result.parsed_rows[0].card_id == "base1-4"


def test_unknown_set_records_row_error(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Set="Made Up Set")])
    result = validate_workbook(blob, db_session)
    assert result.row_errors[0].message == "Set 'Made Up Set' is not recognized"
    assert result.parsed_rows == []


def test_blank_set_required(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Set=None)])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Set is required" in msgs


def test_unknown_card_number(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(**{"Card Number": 999})])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert any("999 does not exist" in m for m in msgs)


def test_blank_card_number_required(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(**{"Card Number": None})])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Card Number is required" in msgs


def test_card_number_float_coerced_to_int_string(db_session, sample_cards):
    """Excel often hands back ``4.0``; we store cards.number as ``"4"``."""
    blob = _build_workbook(rows=[_valid_row(**{"Card Number": 4.0})])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors


def test_card_number_decimal_float_kept_as_decimal_string(db_session, sample_cards):
    """A non-integer float keeps its decimal -- and won't match cards.number."""
    blob = _build_workbook(rows=[_valid_row(**{"Card Number": 4.5})])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert any("does not exist" in m for m in msgs)


def test_card_number_int_left_alone(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(**{"Card Number": 4})])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors


def test_card_number_bool_coerced_to_string(db_session, sample_cards):
    """``True`` is technically an int, but we don't want ``True`` -> ``1``."""
    blob = _build_workbook(rows=[_valid_row(**{"Card Number": True})])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    # 'True' is not a real card number in Base Set, so the lookup fails.
    assert any("does not exist" in m for m in msgs)


def test_card_number_alphanumeric_string(db_session, sample_set):
    """A card with an alphanumeric number still resolves."""
    from models.card import Card

    db_session.add(
        Card(
            id="base1-SV01",
            set_id=sample_set.id,
            name="Special",
            number="SV01",
            rarity="rare",
            supertype="Pokemon",
        )
    )
    db_session.flush()

    blob = _build_workbook(rows=[_valid_row(**{"Card Number": "SV01"})])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors


def test_invalid_condition(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Condition="MINT")])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Condition must be NM, LP, MP, HP, or DMG" in msgs


def test_blank_condition_required(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Condition=None)])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Condition is required" in msgs


def test_condition_lowercase_uppercased(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Condition="nm")])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors


def test_is_first_edition_blank_treated_as_false(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(**{"Is 1st Edition": None})])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors
    assert result.parsed_rows[0].is_first_edition is False


def test_is_first_edition_true(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(**{"Is 1st Edition": "TRUE"})])
    result = validate_workbook(blob, db_session)
    assert result.parsed_rows[0].is_first_edition is True


def test_is_first_edition_invalid(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(**{"Is 1st Edition": "yes"})])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Is 1st Edition must be TRUE or FALSE" in msgs


def test_quantity_blank_required(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Quantity=None)])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Quantity is required" in msgs


def test_quantity_zero_rejected(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Quantity=0)])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Quantity must be a whole number greater than 0" in msgs


def test_quantity_negative_rejected(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Quantity=-3)])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Quantity must be a whole number greater than 0" in msgs


def test_quantity_float_decimal_rejected(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Quantity=1.5)])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Quantity must be a whole number greater than 0" in msgs


def test_quantity_whole_number_float_accepted(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Quantity=2.0)])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors
    assert result.parsed_rows[0].quantity == 2


def test_quantity_string_numeric_accepted(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Quantity="3")])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors
    assert result.parsed_rows[0].quantity == 3


def test_quantity_string_non_numeric_rejected(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Quantity="lots")])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Quantity must be a whole number greater than 0" in msgs


def test_quantity_bool_rejected(db_session, sample_cards):
    """``True`` would otherwise sneak through as int=1 -- block it."""
    blob = _build_workbook(rows=[_valid_row(Quantity=True)])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Quantity must be a whole number greater than 0" in msgs


def test_purchase_price_numeric(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(**{"Purchase Price": 12.50})])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors
    assert str(result.parsed_rows[0].purchase_price) == "12.5"


def test_purchase_price_currency_string_stripped(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(**{"Purchase Price": "$1,234.56"})])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors
    assert str(result.parsed_rows[0].purchase_price) == "1234.56"


def test_purchase_price_invalid_string(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(**{"Purchase Price": "free"})])
    result = validate_workbook(blob, db_session)
    msgs = [e.message for e in result.row_errors]
    assert "Purchase Price must be a number or left blank" in msgs


def test_variant_multi_value_split(db_session, sample_cards):
    blob = _build_workbook(rows=[_valid_row(Variant="Reverse Holo, Misprint")])
    result = validate_workbook(blob, db_session)
    assert not result.has_errors
    assert result.parsed_rows[0].variant == ["Reverse Holo", "Misprint"]


def test_unknown_columns_ignored(db_session, sample_cards):
    """Extra columns (e.g. an Error column from a previous failed upload)
    don't break validation."""
    extra_headers = HEADERS + ("Error",)
    rows = [{**_valid_row(), "Error": "old message"}]
    blob = _build_workbook(rows=rows, headers=extra_headers)
    result = validate_workbook(blob, db_session)
    assert not result.has_errors


# ---------- direct unit tests on private coercion helpers ----------
#
# Excel-via-openpyxl rounds whole-number floats back to ints on read, so
# these branches aren't hit through the workbook-driven tests above. The
# helpers are pure -- exercising them directly is the cheapest way to
# pin down their behavior.


def test_coerce_number_to_text_int_float():
    from services.collection_validator import _coerce_number_to_text

    assert _coerce_number_to_text(4.0) == "4"


def test_coerce_number_to_text_decimal_float():
    from services.collection_validator import _coerce_number_to_text

    assert _coerce_number_to_text(4.5) == "4.5"


def test_coerce_int_whole_number_float():
    from services.collection_validator import _coerce_int

    assert _coerce_int(2.0) == 2


def test_coerce_int_decimal_string():
    """A string like ``"3.5"`` parses as a float but is not an integer."""
    from services.collection_validator import _coerce_int

    assert _coerce_int("3.5") is None
