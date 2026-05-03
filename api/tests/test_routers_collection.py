"""
Tests for routers.collection.

Endpoints exercised end-to-end through the FastAPI TestClient:

* ``GET /collection/template``         -- generates an .xlsx blob.
* ``POST /collection/upload``          -- multipart, success + 422 paths.
* ``POST /collection/upload/annotated``-- annotated workbook download.
* ``POST /collection/mock``            -- bundled-asset upload.
* ``GET /collection/session``          -- read via cookie.
* ``DELETE /collection/session``       -- clear cookie + row.

Cookies in TestClient: by default the underlying httpx client persists
cookies across calls, so a successful ``/upload`` sets the cookie that
the follow-up ``/session`` read uses. Tests that need to bypass this
behaviour clear ``client.cookies`` first.
"""

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import text


HEADERS = (
    "Set",
    "Card Number",
    "Card Name",
    "Condition",
    "Variant",
    "Is 1st Edition",
    "Quantity",
    "Purchase Price",
)


def _build_workbook(rows: list[dict] | None = None, headers: tuple[str, ...] = HEADERS) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Collection"
    for col_idx, label in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=label)
    if rows:
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, label in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(label))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _valid(**overrides):
    base = {
        "Set": "Base Set",
        "Card Number": 4,
        "Condition": "NM",
        "Is 1st Edition": "FALSE",
        "Quantity": 1,
    }
    base.update(overrides)
    return base


@pytest.fixture(autouse=True)
def _insecure_cookies(monkeypatch):
    """TestClient runs over plain HTTP, so a ``Secure`` cookie would be
    silently dropped by httpx. Force the cookie helper to emit
    non-secure cookies for the duration of these tests so the upload
    flow's cookie round-trips into the follow-up session reads."""
    monkeypatch.setenv("SESSION_COOKIE_SECURE", "false")


@pytest.fixture
def session_cleanup(engine):
    """Track session ids the test created and DELETE them at teardown."""
    created: list[str] = []
    yield created
    if created:
        with engine.begin() as conn:
            conn.execute(
                text("DELETE FROM collection_sessions WHERE id = ANY(:ids)"),
                {"ids": created},
            )


# ---------- /collection/template ----------


def test_template_endpoint_returns_xlsx(client, sample_set):
    response = client.get("/collection/template")
    assert response.status_code == 200
    assert response.headers["content-disposition"].startswith("attachment;")
    assert response.headers["content-disposition"].endswith('.xlsx"')
    wb = load_workbook(BytesIO(response.content))
    assert "Collection" in wb.sheetnames
    assert "Instructions" in wb.sheetnames


# ---------- /collection/upload ----------


def test_upload_success_creates_session(client, sample_cards, session_cleanup):
    blob = _build_workbook(rows=[_valid()])
    response = client.post(
        "/collection/upload",
        files={"file": ("collection.xlsx", blob, "application/octet-stream")},
    )
    assert response.status_code == 200
    body = response.json()
    session_cleanup.append(body["session_id"])
    assert body["card_count"] == 1
    assert body["set_count"] == 1
    # Cookie set on the response.
    assert "collection_session_id" in response.cookies


def test_upload_summarizes_quantity_and_set_count(client, sample_cards, session_cleanup):
    rows = [_valid(Quantity=2), _valid(**{"Card Number": 58, "Quantity": 3})]
    blob = _build_workbook(rows=rows)
    response = client.post(
        "/collection/upload",
        files={"file": ("collection.xlsx", blob, "application/octet-stream")},
    )
    assert response.status_code == 200
    body = response.json()
    session_cleanup.append(body["session_id"])
    assert body["card_count"] == 5  # 2 + 3
    assert body["set_count"] == 1


def test_upload_row_errors_returns_422(client, sample_cards):
    blob = _build_workbook(rows=[_valid(Set="Made Up Set")])
    response = client.post(
        "/collection/upload",
        files={"file": ("collection.xlsx", blob, "application/octet-stream")},
    )
    assert response.status_code == 422
    detail = response.json()["detail"]
    assert detail["total_rows"] == 1
    assert len(detail["error_rows"]) == 1
    assert detail["distinct_error_messages"]


def test_upload_structural_error_returns_422(client, sample_cards):
    headers = tuple(h for h in HEADERS if h != "Condition")
    blob = _build_workbook(rows=[_valid()], headers=headers)
    response = client.post(
        "/collection/upload",
        files={"file": ("collection.xlsx", blob, "application/octet-stream")},
    )
    assert response.status_code == 422
    detail = response.json()["detail"]
    assert "Condition" in detail["structural_error"]


# ---------- /collection/upload/annotated ----------


def test_annotated_endpoint_returns_xlsx_with_error_column(client, sample_cards):
    blob = _build_workbook(rows=[_valid(Set="Made Up Set")])
    response = client.post(
        "/collection/upload/annotated",
        files={"file": ("collection.xlsx", blob, "application/octet-stream")},
    )
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    ws = wb["Collection"]
    last_header = ws.cell(row=1, column=ws.max_column).value
    assert last_header == "Error"


# ---------- /collection/mock ----------


def test_mock_endpoint_uses_bundled_file(client, sample_cards, session_cleanup, monkeypatch):
    """Patch MOCK_FILE_PATH to a workbook we control so the test doesn't
    depend on api/assets/mock_collection.xlsx being checked in or up to
    date with the test database fixture."""
    import tempfile
    from pathlib import Path

    from routers import collection as collection_router

    blob = _build_workbook(rows=[_valid()])
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fp:
        fp.write(blob)
        tmp_path = Path(fp.name)
    monkeypatch.setattr(collection_router, "MOCK_FILE_PATH", tmp_path)

    response = client.post("/collection/mock")
    assert response.status_code == 200
    body = response.json()
    session_cleanup.append(body["session_id"])
    assert body["card_count"] == 1


def test_mock_endpoint_returns_503_when_missing(client, monkeypatch):
    from pathlib import Path

    from routers import collection as collection_router

    monkeypatch.setattr(
        collection_router, "MOCK_FILE_PATH", Path("/nonexistent/mock.xlsx")
    )
    response = client.post("/collection/mock")
    assert response.status_code == 503


# ---------- /collection/session ----------


def test_session_get_with_no_cookie_returns_404(client):
    client.cookies.clear()
    response = client.get("/collection/session")
    assert response.status_code == 404


def test_session_get_with_unknown_cookie_returns_404(client):
    client.cookies.clear()
    client.cookies.set("collection_session_id", "no-such-id")
    response = client.get("/collection/session")
    assert response.status_code == 404


def test_session_get_after_upload_returns_rows(client, sample_cards, session_cleanup):
    blob = _build_workbook(rows=[_valid(Quantity=2)])
    upload = client.post(
        "/collection/upload",
        files={"file": ("collection.xlsx", blob, "application/octet-stream")},
    )
    session_cleanup.append(upload.json()["session_id"])

    response = client.get("/collection/session")
    assert response.status_code == 200
    body = response.json()
    assert body["card_count"] == 2
    assert body["set_count"] == 1
    assert len(body["rows"]) == 1
    assert body["rows"][0]["card_id"] == "base1-4"


def test_session_delete_clears_cookie_and_row(client, sample_cards, session_cleanup):
    blob = _build_workbook(rows=[_valid()])
    upload = client.post(
        "/collection/upload",
        files={"file": ("collection.xlsx", blob, "application/octet-stream")},
    )
    session_id = upload.json()["session_id"]
    session_cleanup.append(session_id)

    delete_response = client.delete("/collection/session")
    assert delete_response.status_code == 204

    follow_up = client.get("/collection/session")
    assert follow_up.status_code == 404


def test_session_delete_without_cookie_is_ok(client):
    client.cookies.clear()
    response = client.delete("/collection/session")
    assert response.status_code == 204


# ---------- summarize: cards without dash get unique-set bucket ----------


def test_card_id_without_dash_still_summarized(client, sample_set, session_cleanup, monkeypatch):
    """A card_id that doesn't contain a dash falls into the else-branch of
    ``_summarize`` -- the whole id is treated as the set key."""
    from routers import collection as collection_router
    from schemas.collection import ParsedCollectionRow

    rows = [
        ParsedCollectionRow(
            card_id="weirdcard",
            condition="NM",
            variant=[],
            is_first_edition=False,
            quantity=1,
        ),
    ]
    card_count, set_count = collection_router._summarize(rows)
    assert card_count == 1
    assert set_count == 1


# ---------- cookie Secure flag controlled by env var ----------


def test_cookie_secure_default_true(monkeypatch):
    from routers.collection import _cookie_secure

    monkeypatch.delenv("SESSION_COOKIE_SECURE", raising=False)
    assert _cookie_secure() is True


def test_cookie_secure_false_when_env_says_so(monkeypatch):
    from routers.collection import _cookie_secure

    for value in ("false", "FALSE", "0", "no"):
        monkeypatch.setenv("SESSION_COOKIE_SECURE", value)
        assert _cookie_secure() is False


def test_cookie_secure_true_for_other_values(monkeypatch):
    from routers.collection import _cookie_secure

    monkeypatch.setenv("SESSION_COOKIE_SECURE", "true")
    assert _cookie_secure() is True


# ---------- /collection/cards-with-prices ----------


def test_cards_with_prices_no_session_returns_404(client):
    client.cookies.clear()
    response = client.get("/collection/cards-with-prices")
    assert response.status_code == 404


def test_cards_with_prices_returns_card_metadata(
    client, sample_cards, session_cleanup
):
    blob = _build_workbook(rows=[_valid()])
    upload = client.post(
        "/collection/upload",
        files={"file": ("c.xlsx", blob, "application/octet-stream")},
    )
    session_cleanup.append(upload.json()["session_id"])

    response = client.get("/collection/cards-with-prices")
    assert response.status_code == 200
    body = response.json()
    assert len(body["cards"]) == 1
    card = body["cards"][0]
    assert card["card_id"] == "base1-4"
    assert card["card_name"] == "Charizard"
    assert card["set_name"] == "Base Set"


def test_cards_with_prices_unknown_session_cookie_returns_404(client):
    client.cookies.clear()
    client.cookies.set("collection_session_id", "no-such-id")
    response = client.get("/collection/cards-with-prices")
    assert response.status_code == 404


# ---------- /collection/timeseries ----------


def test_timeseries_no_session_returns_404(client):
    client.cookies.clear()
    response = client.get("/collection/timeseries")
    assert response.status_code == 404


def test_timeseries_returns_points(client, sample_cards, session_cleanup):
    blob = _build_workbook(rows=[_valid()])
    upload = client.post(
        "/collection/upload",
        files={"file": ("c.xlsx", blob, "application/octet-stream")},
    )
    session_cleanup.append(upload.json()["session_id"])

    response = client.get("/collection/timeseries", params={"window": "7d"})
    assert response.status_code == 200
    body = response.json()
    assert "points" in body
    assert "earliest_snapshot" in body


def test_timeseries_invalid_window_returns_422(client, sample_cards, session_cleanup):
    blob = _build_workbook(rows=[_valid()])
    upload = client.post(
        "/collection/upload",
        files={"file": ("c.xlsx", blob, "application/octet-stream")},
    )
    session_cleanup.append(upload.json()["session_id"])

    response = client.get("/collection/timeseries", params={"window": "yearly"})
    assert response.status_code == 422


# ---------- /collection/movers ----------


def test_movers_no_session_returns_404(client):
    client.cookies.clear()
    response = client.get("/collection/movers")
    assert response.status_code == 404


def test_movers_returns_gainers_and_losers_lists(
    client, sample_cards, session_cleanup
):
    blob = _build_workbook(rows=[_valid()])
    upload = client.post(
        "/collection/upload",
        files={"file": ("c.xlsx", blob, "application/octet-stream")},
    )
    session_cleanup.append(upload.json()["session_id"])

    response = client.get(
        "/collection/movers", params={"window": "30d", "count": 5}
    )
    assert response.status_code == 200
    body = response.json()
    assert "gainers" in body
    assert "losers" in body


def test_movers_invalid_window_returns_422(client, sample_cards, session_cleanup):
    blob = _build_workbook(rows=[_valid()])
    upload = client.post(
        "/collection/upload",
        files={"file": ("c.xlsx", blob, "application/octet-stream")},
    )
    session_cleanup.append(upload.json()["session_id"])

    response = client.get("/collection/movers", params={"window": "decade"})
    assert response.status_code == 422


def test_movers_invalid_count_returns_422(client, sample_cards, session_cleanup):
    blob = _build_workbook(rows=[_valid()])
    upload = client.post(
        "/collection/upload",
        files={"file": ("c.xlsx", blob, "application/octet-stream")},
    )
    session_cleanup.append(upload.json()["session_id"])

    response = client.get("/collection/movers", params={"count": 0})
    assert response.status_code == 422


def test_movers_negative_min_pct_returns_422(
    client, sample_cards, session_cleanup
):
    blob = _build_workbook(rows=[_valid()])
    upload = client.post(
        "/collection/upload",
        files={"file": ("c.xlsx", blob, "application/octet-stream")},
    )
    session_cleanup.append(upload.json()["session_id"])

    response = client.get("/collection/movers", params={"min_pct": -0.1})
    assert response.status_code == 422
