"""
Tests for services.collection_excel.

Two layers of coverage:

1. **Row-collection helpers** -- each ``_collect_*`` function is
   exercised directly with fixture-backed snapshots so the SQL,
   filtering, fallback, and ladder logic can be asserted without
   round-tripping through openpyxl.

2. **End-to-end populate_template** -- builds the workbook bytes,
   loads them back through openpyxl, and asserts on table refs / cell
   values / blank handling. The placeholder template is read from
   the real ``api/assets/collection_template.xlsx`` so any drift in
   that asset surfaces here.
"""

from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from schemas.collection import ParsedCollectionRow
from services.collection_excel import (
    EXCEL_MEDIA_TYPE,
    SAMPLE_DAYS,
    TEMPLATE_PATH,
    _collect_details_rows,
    _collect_historic_rows,
    _collect_multipliers_rows,
    _collect_upgrade_rows,
    _locf_price,
    _resolve_market_price,
    _sample_dates,
    excel_filename,
    populate_template,
)


def _row(card_id: str = "base1-4", **overrides) -> ParsedCollectionRow:
    base = dict(
        card_id=card_id,
        condition="NM",
        variant=[],
        is_first_edition=False,
        quantity=1,
        purchase_price=None,
    )
    base.update(overrides)
    return ParsedCollectionRow(**base)


def _add_snapshot(db_session, card_id, captured, market_price, condition="NM", variant=None):
    from models.card import PriceSnapshot

    db_session.add(
        PriceSnapshot(
            card_id=card_id,
            source="tcgplayer",
            condition=condition,
            variant=variant,
            market_price=Decimal(str(market_price)),
            captured_at=datetime.combine(captured, datetime.min.time()),
            captured_date=captured,
        )
    )


# ---------- excel_filename ----------


def test_excel_filename_uses_today_by_default():
    name = excel_filename()
    assert name.startswith("collection-report-")
    assert name.endswith(".xlsx")


def test_excel_filename_uses_provided_date():
    assert excel_filename(date(2026, 5, 3)) == "collection-report-2026-05-03.xlsx"


# ---------- _resolve_market_price ----------


def test_resolve_market_price_direct_hit():
    prices = {("base1-4", "LP"): Decimal("50.00")}
    price, fellback = _resolve_market_price(prices, "base1-4", "LP")
    assert price == Decimal("50.00")
    assert fellback is False


def test_resolve_market_price_fallback_to_nm():
    prices = {("base1-4", "NM"): Decimal("100.00")}
    price, fellback = _resolve_market_price(prices, "base1-4", "LP")
    assert price == Decimal("100.00")
    assert fellback is True


def test_resolve_market_price_no_data():
    price, fellback = _resolve_market_price({}, "base1-4", "LP")
    assert price is None
    assert fellback is False


def test_resolve_market_price_nm_user_no_fallback():
    """A user on NM with no NM data does not fall back to itself."""
    price, fellback = _resolve_market_price({}, "base1-4", "NM")
    assert price is None
    assert fellback is False


# ---------- _sample_dates ----------


def test_sample_dates_returns_twice_monthly_anchors():
    today = date(2026, 5, 20)
    dates = _sample_dates(today, months=2)
    assert dates == [
        date(2026, 4, 1),
        date(2026, 4, 15),
        date(2026, 5, 1),
        date(2026, 5, 15),
    ]


def test_sample_dates_drops_future_anchors():
    today = date(2026, 5, 10)
    dates = _sample_dates(today, months=1)
    # Only April 1, April 15... wait, with months=1 we anchor at May.
    # The 1st has happened; the 15th hasn't.
    assert dates == [date(2026, 5, 1)]


def test_sample_dates_wraps_year_boundary():
    today = date(2026, 2, 28)
    dates = _sample_dates(today, months=3)
    assert dates[0] == date(2025, 12, 1)
    assert dates[-1] == date(2026, 2, 15)


def test_sample_days_constant():
    assert SAMPLE_DAYS == (1, 15)


# ---------- _locf_price ----------


def test_locf_picks_most_recent_at_or_before_target():
    from services.collection_excel import _Snap

    snaps = [
        _Snap(date(2026, 4, 1), Decimal("10")),
        _Snap(date(2026, 4, 15), Decimal("12")),
        _Snap(date(2026, 5, 1), Decimal("15")),
    ]
    assert _locf_price(snaps, date(2026, 4, 20)) == Decimal("12")
    assert _locf_price(snaps, date(2026, 5, 10)) == Decimal("15")


def test_locf_returns_none_when_no_qualifying_snapshot():
    from services.collection_excel import _Snap

    snaps = [_Snap(date(2026, 5, 1), Decimal("10"))]
    assert _locf_price(snaps, date(2026, 4, 1)) is None


# ---------- _collect_details_rows ----------


def test_collect_details_empty_input_returns_empty(db_session):
    assert _collect_details_rows(db_session, []) == []


def test_collect_details_populates_metadata_and_market_price(
    db_session, sample_cards
):
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "120.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", quantity=2, purchase_price=Decimal("100"))]
    [row] = _collect_details_rows(db_session, rows)
    assert row["card_name"] == "Charizard"
    assert row["set_name"] == "Base Set"
    assert row["market_price"] == Decimal("120.00")
    assert row["total_value"] == Decimal("240.00")
    assert row["gain_dollar"] == Decimal("40.00")
    assert row["gain_percent"] == Decimal("0.2")
    assert row["pricing_warning"] is False


def test_collect_details_falls_back_to_nm_with_warning(db_session, sample_cards):
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "120.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", condition="LP")]
    [row] = _collect_details_rows(db_session, rows)
    assert row["market_price"] == Decimal("120.00")  # NM fallback
    assert row["pricing_warning"] is True


def test_collect_details_marks_warning_for_variant_cards(
    db_session, sample_cards
):
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "120.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", variant=["Reverse Holo"])]
    [row] = _collect_details_rows(db_session, rows)
    assert row["pricing_warning"] is True
    assert row["variant"] == "Reverse Holo"


def test_collect_details_joins_multiple_variants_with_comma(
    db_session, sample_cards
):
    rows = [_row("base1-4", variant=["Reverse Holo", "Misprint"])]
    [row] = _collect_details_rows(db_session, rows)
    assert row["variant"] == "Reverse Holo, Misprint"


def test_collect_details_blanks_for_missing_purchase_price(
    db_session, sample_cards
):
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "120.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", purchase_price=None)]
    [row] = _collect_details_rows(db_session, rows)
    assert row["purchase_price"] is None
    assert row["gain_dollar"] is None
    assert row["gain_percent"] is None


def test_collect_details_no_market_data_blanks_value_fields(
    db_session, sample_cards
):
    rows = [_row("base1-4", purchase_price=Decimal("100"))]
    [row] = _collect_details_rows(db_session, rows)
    assert row["market_price"] is None
    assert row["total_value"] is None
    assert row["gain_dollar"] is None
    assert row["gain_percent"] is None
    assert row["pricing_warning"] is False  # no fallback occurred, no variant


def test_collect_details_zero_purchase_price_yields_no_percent(
    db_session, sample_cards
):
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "120.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", purchase_price=Decimal("0"))]
    [row] = _collect_details_rows(db_session, rows)
    assert row["gain_dollar"] == Decimal("120.00")
    assert row["gain_percent"] is None  # divide-by-zero guarded


def test_collect_details_set_total_with_secrets_uses_total_count(
    db_session, sample_cards
):
    """The Excel column maps to total_count (M03_S07's substitution)."""
    rows = [_row("base1-4")]
    [row] = _collect_details_rows(db_session, rows)
    # Two cards in fixture, so total_count = 2.
    assert row["set_total_with_secrets"] == 2


# ---------- _collect_multipliers_rows ----------


def test_collect_multipliers_empty_input_returns_empty(db_session):
    assert _collect_multipliers_rows(db_session, []) == []


def test_collect_multipliers_filters_to_user_sets(
    db_session, sample_set, sample_multipliers
):
    rows = _collect_multipliers_rows(db_session, [sample_set.id])
    assert all(r["set_id"] == sample_set.id for r in rows)
    # Both grouping types appear (rarity + supertype) per spec.
    assert {r["grouping_type"] for r in rows} == {"rarity", "supertype"}


def test_collect_multipliers_excludes_other_sets(
    db_session, sample_set, sample_multipliers
):
    rows = _collect_multipliers_rows(db_session, ["nonexistent_set"])
    assert rows == []


# ---------- _collect_historic_rows ----------


def test_collect_historic_empty_input_returns_empty(db_session):
    assert _collect_historic_rows(db_session, []) == []


def test_collect_historic_skips_graded_conditions(db_session, sample_cards):
    rows = [_row("base1-4", condition="PSA-10")]
    assert _collect_historic_rows(db_session, rows) == []


def test_collect_historic_returns_one_row_per_sample_with_data(
    db_session, sample_cards
):
    today = date.today()
    # Single snapshot at the start of the window -- LOCF carries it forward.
    _add_snapshot(db_session, "base1-4", today - timedelta(days=180), "10.00")
    db_session.flush()
    out = _collect_historic_rows(db_session, [_row("base1-4")])
    assert len(out) > 0
    assert all(row["card_id"] == "base1-4" for row in out)
    assert all(row["condition"] == "NM" for row in out)
    # Sample dates are exclusively the 1st or 15th of a month.
    assert all(row["sample_date"].day in (1, 15) for row in out)


def test_collect_historic_skips_samples_before_earliest_snapshot(
    db_session, sample_cards
):
    today = date.today()
    # Snapshot three days ago -- only the most recent sample dates will
    # have qualifying snapshots; earlier anchors are skipped (no NULL).
    _add_snapshot(db_session, "base1-4", today - timedelta(days=3), "10.00")
    db_session.flush()
    out = _collect_historic_rows(db_session, [_row("base1-4")])
    # Either 0 or 1 row depending on whether today is past the next sample
    # anchor, but importantly all rows have a real price.
    assert all(row["market_price"] is not None for row in out)


# ---------- _collect_upgrade_rows ----------


def test_collect_upgrade_empty_input_returns_empty(db_session):
    assert _collect_upgrade_rows(db_session, []) == []


def test_collect_upgrade_skips_nm_cards(db_session, sample_cards):
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "120.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", condition="NM")]
    assert _collect_upgrade_rows(db_session, rows) == []


def test_collect_upgrade_skips_graded_cards(db_session, sample_cards):
    rows = [_row("base1-4", condition="PSA-10")]
    assert _collect_upgrade_rows(db_session, rows) == []


def test_collect_upgrade_emits_one_row_per_target_with_data(
    db_session, sample_cards
):
    today = date.today()
    # User on LP -> only NM is an upgrade target. Seed both.
    _add_snapshot(db_session, "base1-4", today, "100.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", condition="LP")]
    out = _collect_upgrade_rows(db_session, rows)
    assert len(out) == 1
    assert out[0]["condition"] == "NM"
    assert out[0]["market_price"] == Decimal("100.00")


def test_collect_upgrade_emits_full_ladder_for_dmg(db_session, sample_cards):
    today = date.today()
    for cond, price in [("NM", "100"), ("LP", "70"), ("MP", "50"), ("HP", "30")]:
        _add_snapshot(db_session, "base1-4", today, price, condition=cond)
    db_session.flush()
    rows = [_row("base1-4", condition="DMG")]
    out = _collect_upgrade_rows(db_session, rows)
    conditions = sorted([r["condition"] for r in out])
    assert conditions == ["HP", "LP", "MP", "NM"]


def test_collect_upgrade_skips_target_with_no_price(db_session, sample_cards):
    today = date.today()
    # User on MP. Only NM has data -- LP target is skipped (no row, not null).
    _add_snapshot(db_session, "base1-4", today, "100.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", condition="MP")]
    out = _collect_upgrade_rows(db_session, rows)
    assert [r["condition"] for r in out] == ["NM"]


def test_collect_upgrade_dedupes_across_session_rows(
    db_session, sample_cards
):
    """Two session rows for the same (card, condition) only produce one
    row per upgrade target."""
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "100.00", condition="NM")
    db_session.flush()
    rows = [
        _row("base1-4", condition="LP", quantity=1),
        _row("base1-4", condition="LP", quantity=2),
    ]
    out = _collect_upgrade_rows(db_session, rows)
    assert len(out) == 1


def test_collect_upgrade_dedupes_overlapping_targets(db_session, sample_cards):
    """When the same card appears at multiple conditions whose upgrade
    targets overlap (LP and MP both reach NM), the overlapping target
    only contributes one row."""
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "100.00", condition="NM")
    _add_snapshot(db_session, "base1-4", today, "70.00", condition="LP")
    db_session.flush()
    rows = [
        _row("base1-4", condition="LP"),  # target: NM
        _row("base1-4", condition="MP"),  # targets: LP, NM
    ]
    out = _collect_upgrade_rows(db_session, rows)
    target_set = sorted([r["condition"] for r in out])
    # NM appears once even though both session rows reach it.
    assert target_set == ["LP", "NM"]


# ---------- populate_template (end-to-end) ----------


def test_populate_template_writes_all_four_tables(db_session, sample_cards):
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "120.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", quantity=2)]

    blob = populate_template(db_session, rows)
    wb = load_workbook(BytesIO(blob))
    table_names = {
        name for sheet in wb.worksheets for name in sheet.tables
    }
    # The template may carry additional Tables that ship with it
    # (e.g. Power Query output tables). We only assert that the four
    # we populate are present.
    assert {
        "collection_details",
        "condition_multipliers",
        "historic_prices",
        "card_prices_all_conditions",
    }.issubset(table_names)


def test_populate_template_preserves_template_structure(
    db_session, sample_cards
):
    blob = populate_template(db_session, [_row("base1-4")])
    wb = load_workbook(BytesIO(blob))
    sheet_names = {s.title for s in wb.worksheets}
    assert {
        "Collection Details",
        "Condition Multipliers",
        "Historic Prices",
        "Card Prices",
    }.issubset(sheet_names)


def test_populate_template_writes_real_booleans(db_session, sample_cards):
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "120.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", is_first_edition=True)]

    blob = populate_template(db_session, rows)
    wb = load_workbook(BytesIO(blob))
    sheet = wb["Collection Details"]
    headers = [c.value for c in sheet[1]]
    is_first_idx = headers.index("is_first_edition") + 1
    assert sheet.cell(row=2, column=is_first_idx).value is True


def test_populate_template_blank_for_missing_purchase_price(
    db_session, sample_cards
):
    today = date.today()
    _add_snapshot(db_session, "base1-4", today, "120.00", condition="NM")
    db_session.flush()
    rows = [_row("base1-4", purchase_price=None)]

    blob = populate_template(db_session, rows)
    wb = load_workbook(BytesIO(blob))
    sheet = wb["Collection Details"]
    headers = [c.value for c in sheet[1]]
    purchase_idx = headers.index("purchase_price") + 1
    gain_dollar_idx = headers.index("gain_dollar") + 1
    gain_percent_idx = headers.index("gain_percent") + 1
    assert sheet.cell(row=2, column=purchase_idx).value is None
    assert sheet.cell(row=2, column=gain_dollar_idx).value is None
    assert sheet.cell(row=2, column=gain_percent_idx).value is None


def test_populate_template_empty_session_keeps_placeholder_refs(
    db_session,
):
    """An empty session still produces a valid workbook -- each table
    keeps its 2-row placeholder ref so Excel doesn't reject it."""
    blob = populate_template(db_session, [])
    wb = load_workbook(BytesIO(blob))
    for sheet in wb.worksheets:
        for tname in list(sheet.tables):
            tab = sheet.tables[tname]
            assert tab.ref.endswith(":") is False  # well-formed
            # ref should still terminate at row 2 (header + one blank row)
            last_row = int("".join(c for c in tab.ref.split(":")[1] if c.isdigit()))
            assert last_row == 2


def test_populate_template_extends_table_ref_with_data(
    db_session, sample_cards
):
    """The ``collection_details`` Table ref must extend to cover the
    populated rows -- otherwise Excel would only consider the placeholder
    range."""
    rows = [_row("base1-4"), _row("base1-58")]
    blob = populate_template(db_session, rows)
    wb = load_workbook(BytesIO(blob))
    sheet = wb["Collection Details"]
    tab = sheet.tables["collection_details"]
    last_row = int("".join(c for c in tab.ref.split(":")[1] if c.isdigit()))
    assert last_row == 3  # header + 2 data rows


def test_populate_template_clears_prior_data_rows(db_session, sample_cards):
    """The populator must overwrite any leftover rows from a previous
    run -- otherwise stale data would leak between users / sessions."""
    # First run with two rows; second run with one row should not leave
    # the second card's data hanging around.
    blob_first = populate_template(
        db_session, [_row("base1-4"), _row("base1-58")]
    )
    blob_second = populate_template(db_session, [_row("base1-4")])
    wb = load_workbook(BytesIO(blob_second))
    sheet = wb["Collection Details"]
    headers = [c.value for c in sheet[1]]
    card_id_idx = headers.index("card_id") + 1
    # Row 2 has data; row 3 should be entirely cleared.
    assert sheet.cell(row=2, column=card_id_idx).value == "base1-4"
    assert sheet.cell(row=3, column=card_id_idx).value is None


def test_populate_template_missing_template_raises(monkeypatch, db_session):
    from services import collection_excel as mod
    from pathlib import Path

    monkeypatch.setattr(mod, "TEMPLATE_PATH", Path("/nonexistent/path.xlsx"))
    import pytest

    with pytest.raises(FileNotFoundError):
        populate_template(db_session, [_row("base1-4")])


def test_template_constants_are_accessible():
    """Smoke-check the public surface: file path + media type live in
    the module so the router imports stay tidy."""
    assert TEMPLATE_PATH.name == "collection_template.xlsx"
    assert "spreadsheetml.sheet" in EXCEL_MEDIA_TYPE
