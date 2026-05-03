"""
Tests for services.collection_annotator.

The annotator re-runs validation, then writes an Error column + red
fill into a copy of the input workbook. Tests build minimal workbooks
in-memory and assert on the parsed output rather than raw bytes.
"""

from io import BytesIO

from openpyxl import Workbook, load_workbook

from services.collection_annotator import ERROR_HEADER, annotate_workbook


HEADERS = (
    "Set",
    "Card Number",
    "Card Name",
    "Condition",
    "Variant",
    "Is 1st Edition",
    "Quantity",
    "Purchase Price",
)


def _build(rows: list[dict] | None = None, headers: tuple[str, ...] = HEADERS) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Collection"
    for col_idx, label in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=label)
    if rows:
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, label in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(label))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _valid(**overrides):
    base = {
        "Set": "Base Set",
        "Card Number": 4,
        "Condition": "NM",
        "Is 1st Edition": "FALSE",
        "Quantity": 1,
    }
    base.update(overrides)
    return base


def test_annotator_adds_error_column(db_session, sample_cards):
    blob = _build(rows=[_valid(Set="Made Up Set")])
    annotated = annotate_workbook(blob, db_session)
    wb = load_workbook(BytesIO(annotated))
    ws = wb["Collection"]
    last_header = ws.cell(row=1, column=ws.max_column).value
    assert last_header == ERROR_HEADER


def test_annotator_writes_error_message_for_invalid_row(db_session, sample_cards):
    blob = _build(rows=[_valid(Set="Made Up Set")])
    annotated = annotate_workbook(blob, db_session)
    wb = load_workbook(BytesIO(annotated))
    ws = wb["Collection"]
    error_col = ws.max_column
    msg = ws.cell(row=2, column=error_col).value
    assert "Made Up Set" in msg


def test_annotator_leaves_valid_rows_blank(db_session, sample_cards):
    """Valid rows have a blank Error cell so re-uploading the annotated file
    after a fix succeeds without the user having to delete the column."""
    blob = _build(rows=[_valid(), _valid(Set="Made Up Set")])
    annotated = annotate_workbook(blob, db_session)
    wb = load_workbook(BytesIO(annotated))
    ws = wb["Collection"]
    error_col = ws.max_column
    assert ws.cell(row=2, column=error_col).value in (None, "")
    assert ws.cell(row=3, column=error_col).value is not None


def test_annotator_joins_multiple_messages_per_row(db_session, sample_cards):
    """Two errors on one row -> joined with '; ' in the Error cell."""
    blob = _build(rows=[_valid(Set="Made Up Set", Condition="MINT")])
    annotated = annotate_workbook(blob, db_session)
    wb = load_workbook(BytesIO(annotated))
    ws = wb["Collection"]
    error_col = ws.max_column
    msg = ws.cell(row=2, column=error_col).value
    assert ";" in msg


def test_annotator_fills_invalid_rows_red(db_session, sample_cards):
    blob = _build(rows=[_valid(Set="Made Up Set")])
    annotated = annotate_workbook(blob, db_session)
    wb = load_workbook(BytesIO(annotated))
    ws = wb["Collection"]
    fill = ws.cell(row=2, column=1).fill
    assert fill.fgColor.value.endswith("5C2828")


def test_annotator_reuses_existing_error_column(db_session, sample_cards):
    """An ``Error`` header already in the workbook is reused, not duplicated."""
    headers = HEADERS + ("Error",)
    blob = _build(rows=[_valid(Set="Made Up Set")], headers=headers)
    annotated = annotate_workbook(blob, db_session)
    wb = load_workbook(BytesIO(annotated))
    ws = wb["Collection"]
    # max_column should still match the original headers count (no new column).
    assert ws.max_column == len(headers)
    error_col = len(headers)
    assert ws.cell(row=1, column=error_col).value == ERROR_HEADER


def test_annotator_handles_structural_error(db_session, sample_cards):
    """Missing required column -> the structural message lands on row 2 col Error."""
    headers = tuple(h for h in HEADERS if h != "Condition")
    blob = _build(rows=[_valid()], headers=headers)
    annotated = annotate_workbook(blob, db_session)
    wb = load_workbook(BytesIO(annotated))
    ws = wb["Collection"]
    error_col = ws.max_column
    assert ws.cell(row=1, column=error_col).value == ERROR_HEADER
    msg = ws.cell(row=2, column=error_col).value
    assert "Condition" in msg


def test_annotator_no_errors_leaves_rows_clean(db_session, sample_cards):
    """All-valid workbook gets the Error column added but no fills/messages."""
    blob = _build(rows=[_valid()])
    annotated = annotate_workbook(blob, db_session)
    wb = load_workbook(BytesIO(annotated))
    ws = wb["Collection"]
    error_col = ws.max_column
    assert ws.cell(row=2, column=error_col).value in (None, "")
    fill = ws.cell(row=2, column=1).fill
    # Default fill has no fg color or "00000000".
    fg = fill.fgColor.value if fill.fgColor else None
    assert fg in (None, "00000000")
