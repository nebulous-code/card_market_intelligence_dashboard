"""
Tests for services.collection_template.

The template generator queries the live ``sets`` table for the dropdown,
so tests rely on the existing ``sample_set`` fixture. We assert on the
parsed output -- header row, dropdown formula, defined name -- rather
than on raw bytes, so changes to openpyxl's compression don't break the
tests.
"""

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from services.collection_template import (
    HEADERS,
    build_template_workbook,
    template_filename,
)


def test_template_filename_includes_today_by_default():
    name = template_filename()
    assert name.startswith("card-collection-template-")
    assert name.endswith(".xlsx")


def test_template_filename_uses_provided_date():
    assert (
        template_filename(date(2026, 5, 3))
        == "card-collection-template-2026-05-03.xlsx"
    )


def test_template_has_expected_headers(db_session, sample_set):
    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Collection"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert headers == [h for h, _ in HEADERS]


def test_template_freezes_header_row(db_session, sample_set):
    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Collection"]
    assert ws.freeze_panes == "A2"


def test_template_lists_sheet_holds_set_names(db_session, sample_set):
    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    assert "_lists" in wb.sheetnames
    lists_ws = wb["_lists"]
    set_names = [row[0].value for row in lists_ws.iter_rows()]
    assert sample_set.name in set_names
    # Hidden sheet -- not visible to the user in Excel by default.
    assert lists_ws.sheet_state == "hidden"


def test_template_defines_setnames_range(db_session, sample_set):
    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    assert "SetNames" in wb.defined_names
    ref = wb.defined_names["SetNames"].attr_text
    assert ref.startswith("_lists!$A$1:$A$")


def test_template_no_defined_name_when_no_sets(db_session):
    """No sets -> no defined name -> Set dropdown is silently skipped."""
    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    assert "SetNames" not in wb.defined_names


def test_template_includes_condition_dropdown(db_session, sample_set):
    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Collection"]
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert any("NM,LP,MP,HP,DMG" in f for f in formulas if f)


def test_template_includes_first_edition_dropdown(db_session, sample_set):
    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Collection"]
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert any("TRUE,FALSE" in f for f in formulas if f)


def test_template_includes_set_dropdown_when_sets_exist(db_session, sample_set):
    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Collection"]
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert any("SetNames" in f for f in formulas if f)


def test_template_skips_set_dropdown_when_no_sets(db_session):
    """An empty sets table -> no Set DV is added."""
    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    ws = wb["Collection"]
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert not any("SetNames" in f for f in formulas if f)


def test_template_instructions_sheet_present(db_session, sample_set):
    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    assert "Instructions" in wb.sheetnames
    ws = wb["Instructions"]
    text = "\n".join(
        c.value for c in ws["A"] if c.value is not None
    )
    assert "How to use this template" in text
    assert "Sample row" in text


def test_template_query_skips_falsy_set_names(db_session, sample_set):
    """An empty-string set name (defensive) is dropped from the dropdown."""
    from datetime import date

    from models.set import Set

    db_session.add(
        Set(
            id="empty_name",
            name="",
            series="X",
            printed_total=0,
            release_date=date(2000, 1, 1),
        )
    )
    db_session.flush()

    blob = build_template_workbook(db_session)
    wb = load_workbook(BytesIO(blob))
    lists_ws = wb["_lists"]
    set_names = [row[0].value for row in lists_ws.iter_rows()]
    assert "" not in set_names
    assert sample_set.name in set_names
